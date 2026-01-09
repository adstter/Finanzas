"""
Actualizar Datos del Presupuesto 2026
=====================================
Este script extrae los datos del Excel y actualiza el archivo datos-presupuesto.js
para que la version compartida (presupuesto-viewer.html) muestre los datos actualizados.

Uso:
    python actualizar-datos.py

Luego hacer commit y push a GitHub:
    git add datos-presupuesto.js
    git commit -m "Actualizar datos presupuesto"
    git push
"""

import json
import os
import subprocess
import re

MESES = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    import openpyxl

# Rutas
EXCEL_PATH = os.path.expanduser("~/Downloads/Presupuesto 2026.xlsx")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "datos-presupuesto.js")

def get_mensual(sheet, row_num):
    """Extrae valores mensuales (columnas D-O)"""
    return [float(sheet.cell(row=row_num, column=c).value or 0) for c in range(4, 16)]

def get_total(sheet, row_num):
    """Extrae total anual (columna P)"""
    return float(sheet.cell(row=row_num, column=16).value or 0)

def cargar_ventas_actuales():
    """Carga las ventas de productos actuales desde datos-presupuesto.js"""
    try:
        with open(OUTPUT_PATH, 'r', encoding='utf-8') as f:
            contenido = f.read()
            # Buscar ventasProductos en el archivo
            match = re.search(r'"ventasProductos":\s*\[([^\]]+)\]', contenido)
            if match:
                valores = match.group(1).split(',')
                return [float(v.strip()) for v in valores]
    except:
        pass
    return [0] * 12

def solicitar_ventas():
    """Solicita al usuario los valores de ventas de productos"""
    print("\n" + "=" * 50)
    print("VENTAS DE PRODUCTOS PROYECTADAS")
    print("=" * 50)

    ventas_actuales = cargar_ventas_actuales()
    total_actual = sum(ventas_actuales)

    print(f"\nValores actuales de ventas:")
    for i, mes in enumerate(MESES):
        print(f"  {mes}: Q {ventas_actuales[i]:,.2f}")
    print(f"  Total Anual: Q {total_actual:,.2f}")

    print("\nOpciones:")
    print("  1. Mantener valores actuales")
    print("  2. Establecer un valor igual para todos los meses")
    print("  3. Ingresar valor por cada mes")

    opcion = input("\nSelecciona una opcion (1/2/3): ").strip()

    if opcion == '1':
        print("Manteniendo valores actuales.")
        return ventas_actuales

    elif opcion == '2':
        while True:
            try:
                valor = input("\nIngresa el valor mensual para todos los meses: Q ").strip().replace(',', '')
                valor = float(valor)
                ventas = [valor] * 12
                print(f"\nTotal anual: Q {sum(ventas):,.2f}")
                confirmar = input("Confirmar? (s/n): ").strip().lower()
                if confirmar == 's':
                    return ventas
            except ValueError:
                print("Por favor ingresa un numero valido.")

    elif opcion == '3':
        ventas = []
        print("\nIngresa el valor de ventas para cada mes:")
        for i, mes in enumerate(MESES):
            while True:
                try:
                    valor_default = ventas_actuales[i]
                    valor = input(f"  {mes} (actual: Q {valor_default:,.2f}): Q ").strip().replace(',', '')
                    if valor == '':
                        valor = valor_default
                    else:
                        valor = float(valor)
                    ventas.append(valor)
                    break
                except ValueError:
                    print("    Por favor ingresa un numero valido.")
        print(f"\nTotal anual: Q {sum(ventas):,.2f}")
        confirmar = input("Confirmar? (s/n): ").strip().lower()
        if confirmar == 's':
            return ventas
        else:
            return ventas_actuales

    return ventas_actuales

def main():
    print("=" * 50)
    print("ACTUALIZADOR DE DATOS - PRESUPUESTO 2026")
    print("=" * 50)

    # Verificar que existe el Excel
    if not os.path.exists(EXCEL_PATH):
        print(f"\nError: No se encontro el archivo Excel en:")
        print(f"  {EXCEL_PATH}")
        print("\nAsegurate de que el archivo 'Presupuesto 2026.xlsx' este en tu carpeta Downloads.")
        input("\nPresiona Enter para salir...")
        return

    print(f"\nLeyendo: {EXCEL_PATH}")

    # Cargar Excel
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet = wb['Presupuesto 2026']

    print("Extrayendo datos...")

    # Extraer clientes
    clientes = []
    for i in range(3, 51):
        nombre = sheet.cell(row=i, column=2).value
        if nombre and isinstance(nombre, str) and nombre.strip():
            nombre = nombre.strip()
            if any(x in nombre.lower() for x in ['ingresos', 'ventas', 'total', 'reduccion']):
                continue
            mensual = get_mensual(sheet, i)
            total = get_total(sheet, i)
            if total > 0:
                clientes.append({
                    'nombre': nombre,
                    'mensual': mensual,
                    'total': total
                })

    clientes.sort(key=lambda x: x['total'], reverse=True)
    print(f"  - {len(clientes)} clientes encontrados")

    # Estructura de datos
    datos = {
        'ingresos': {
            'clientes': clientes,
            'totalMensual': get_mensual(sheet, 52),
            'totalAnual': get_total(sheet, 52)
        },
        'costos': {
            'detalle': [],
            'totalMensual': get_mensual(sheet, 68),
            'totalAnual': get_total(sheet, 68)
        },
        'margenBruto': {
            'mensual': get_mensual(sheet, 70),
            'anual': get_total(sheet, 70),
            'porcentaje': 0
        },
        'opex': {
            'consultores': [],
            'software': [],
            'otros': [],
            'totalMensual': get_mensual(sheet, 123),
            'totalAnual': get_total(sheet, 123)
        },
        'resultado': {
            'mensual': get_mensual(sheet, 152),
            'anual': get_total(sheet, 152),
            'margenOperativo': 0
        },
        'gastos': []
    }

    # Costos detalle
    for i, nombre_default in [(62, 'Comisiones Recurrentes'), (63, 'Comisiones Gerencia'), (64, 'Costes Servidores')]:
        nombre = sheet.cell(row=i, column=2).value or nombre_default
        mensual = get_mensual(sheet, i)
        total = get_total(sheet, i)
        if total > 0:
            datos['costos']['detalle'].append({
                'nombre': str(nombre),
                'mensual': mensual,
                'total': total
            })
    print(f"  - {len(datos['costos']['detalle'])} costos directos")

    # OPEX Consultores
    for i in range(73, 80):
        nombre = sheet.cell(row=i, column=2).value
        if nombre and isinstance(nombre, str):
            total = get_total(sheet, i)
            if total > 0:
                datos['opex']['consultores'].append({
                    'nombre': nombre.strip(),
                    'categoria': 'Consultores',
                    'mensual': get_mensual(sheet, i),
                    'total': total
                })
    print(f"  - {len(datos['opex']['consultores'])} consultores")

    # OPEX Software
    for i in range(89, 106):
        nombre = sheet.cell(row=i, column=2).value
        if nombre and isinstance(nombre, str):
            total = get_total(sheet, i)
            if total > 0:
                datos['opex']['software'].append({
                    'nombre': nombre.strip(),
                    'categoria': 'Software',
                    'mensual': get_mensual(sheet, i),
                    'total': total
                })
    print(f"  - {len(datos['opex']['software'])} software/servicios")

    # OPEX Otros
    for i in range(111, 120):
        nombre = sheet.cell(row=i, column=2).value
        if nombre and isinstance(nombre, str):
            total = get_total(sheet, i)
            if total > 0:
                datos['opex']['otros'].append({
                    'nombre': nombre.strip(),
                    'categoria': 'Otros',
                    'mensual': get_mensual(sheet, i),
                    'total': total
                })
    print(f"  - {len(datos['opex']['otros'])} otros gastos")

    # Calcular porcentajes
    if datos['ingresos']['totalAnual'] > 0:
        datos['margenBruto']['porcentaje'] = datos['margenBruto']['anual'] / datos['ingresos']['totalAnual'] * 100
        datos['resultado']['margenOperativo'] = datos['resultado']['anual'] / datos['ingresos']['totalAnual'] * 100

    # Consolidar gastos para ranking
    datos['gastos'] = (
        datos['costos']['detalle'] +
        datos['opex']['consultores'] +
        datos['opex']['software'] +
        datos['opex']['otros']
    )
    datos['gastos'].sort(key=lambda x: x['total'], reverse=True)

    # Solicitar ventas de productos
    ventas_productos = solicitar_ventas()
    datos['ventasProductos'] = ventas_productos
    datos['margenProductos'] = 0.30  # 30% margen bruto

    # Guardar archivo JS
    print(f"\nGuardando: {OUTPUT_PATH}")
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write('// Datos del Presupuesto 2026 - Generado automaticamente\n')
        f.write('// Ultima actualizacion: ' + __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S') + '\n')
        f.write('const DATOS_PRESUPUESTO = ')
        f.write(json.dumps(datos, ensure_ascii=False, indent=2))
        f.write(';\n')

    print("\n" + "=" * 50)
    print("DATOS ACTUALIZADOS EXITOSAMENTE")
    print("=" * 50)
    print(f"\nResumen Ingresos Recurrentes:")
    print(f"  Ingresos Totales: Q {datos['ingresos']['totalAnual']:,.2f}")
    print(f"  Costos Directos:  Q {datos['costos']['totalAnual']:,.2f}")
    print(f"  Margen Bruto:     {datos['margenBruto']['porcentaje']:.1f}%")
    print(f"  OPEX Total:       Q {datos['opex']['totalAnual']:,.2f}")
    print(f"  Resultado Neto:   Q {datos['resultado']['anual']:,.2f}")

    total_ventas = sum(datos['ventasProductos'])
    margen_ventas = total_ventas * datos['margenProductos']
    print(f"\nVentas de Productos:")
    print(f"  Ventas Totales:   Q {total_ventas:,.2f}")
    print(f"  Margen (30%):     Q {margen_ventas:,.2f}")

    print(f"\nConsolidado:")
    print(f"  Ingresos Totales: Q {datos['ingresos']['totalAnual'] + total_ventas:,.2f}")
    print(f"  Resultado Final:  Q {datos['resultado']['anual'] + margen_ventas:,.2f}")

    print("\n" + "-" * 50)
    print("Para publicar los cambios, ejecuta:")
    print("  git add datos-presupuesto.js")
    print('  git commit -m "Actualizar datos presupuesto"')
    print("  git push")
    print("-" * 50)

    # Preguntar si quiere hacer push automatico
    respuesta = input("\nDeseas hacer push a GitHub ahora? (s/n): ").strip().lower()
    if respuesta == 's':
        print("\nSubiendo a GitHub...")
        try:
            os.chdir(os.path.dirname(__file__))
            subprocess.run(['git', 'add', 'datos-presupuesto.js'], check=True)
            subprocess.run(['git', 'commit', '-m', 'Actualizar datos presupuesto'], check=True)
            subprocess.run(['git', 'push'], check=True)
            print("\nCambios publicados exitosamente!")
            print("El link compartido se actualizara en 1-2 minutos.")
        except subprocess.CalledProcessError as e:
            print(f"\nError al hacer push: {e}")
            print("Intenta manualmente con los comandos de arriba.")

    input("\nPresiona Enter para salir...")

if __name__ == '__main__':
    main()
